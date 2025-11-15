"""
Utility to extract reference data from Excel file for test fixtures.

This script reads the current state of PRELIM_v1.6.xlsm and extracts:
- Input parameters (assay selection, configuration)
- Calculated results (yields, energy, emissions)
- Creates a JSON test fixture for validation

Usage:
    python extract_reference_data.py --output tc_001_current.json
"""

import openpyxl
import json
from datetime import datetime
from typing import Dict, Any
import sys
import os

# Add parent directory to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))


class ExcelReferenceExtractor:
    """Extract reference data from PRELIM Excel file."""
    
    def __init__(self, excel_path: str):
        """
        Initialize extractor.
        
        Args:
            excel_path: Path to PRELIM Excel file
        """
        self.excel_path = excel_path
        self.wb_values = openpyxl.load_workbook(excel_path, data_only=True)
    
    def extract_current_assay(self) -> Dict[str, Any]:
        """Extract currently selected assay information."""
        try:
            # The assay is selected from Assay Inventory
            # Main Input & Output shows properties from the selected assay
            main_sheet = self.wb_values['Main Input & Output']
            
            # Try to get assay name/description from display area
            # Row 19-20 typically show the assay info
            assay_info = main_sheet['E19'].value or main_sheet['E20'].value
            
            return {
                'assay_info': str(assay_info) if assay_info else 'Unknown',
                'source_sheet': 'Main Input & Output',
                'note': 'Assay identification from displayed properties'
            }
        except Exception as e:
            return {'error': str(e)}
    
    def extract_product_yields(self) -> Dict[str, float]:
        """Extract product yields from Results Single Assay sheet."""
        results = {}
        
        try:
            sheet = self.wb_values['Results Single Assay']
            
            # Product yields in volume % (Column C, starting around row 13)
            products = {
                'gasoline': 'C13',
                'jet': 'C14',
                'diesel': 'C15',
                'fuel_oil': 'C16',
                'coke': 'C17',
                'liquid_heavy_ends': 'C18',
                'rfg': 'C20',
                'lpg': 'C22',
                'petrochemical_feedstocks': 'C23',
                'asphalt': 'C24',
                'atres': 'C25',
                'lube': 'C26',
                'btx': 'C27',
                'naphtha': 'C28',
                'vgo': 'C29'
            }
            
            for product, cell in products.items():
                value = sheet[cell].value
                if value is not None and isinstance(value, (int, float)):
                    results[f'{product}_vol_pct'] = float(value)
            
            # Product flows in bpd (Column D)
            flows = {
                'gasoline': 'D13',
                'jet': 'D14',
                'diesel': 'D15',
                'fuel_oil': 'D16',
                'coke': 'D17',
            }
            
            for product, cell in flows.items():
                value = sheet[cell].value
                if value is not None and isinstance(value, (int, float)):
                    results[f'{product}_bpd'] = float(value)
            
        except Exception as e:
            results['extraction_error'] = str(e)
        
        return results
    
    def extract_energy_data(self) -> Dict[str, float]:
        """Extract energy consumption data."""
        results = {}
        
        try:
            sheet = self.wb_values['Results Single Assay']
            
            # Energy data typically in later rows
            # We'll need to scan for it
            # For now, placeholder
            results['note'] = 'Energy extraction needs cell location mapping'
            
        except Exception as e:
            results['error'] = str(e)
        
        return results
    
    def extract_emissions_data(self) -> Dict[str, float]:
        """Extract emissions data."""
        results = {}
        
        try:
            sheet = self.wb_values['Results Single Assay']
            
            # Emissions data location needs mapping
            results['note'] = 'Emissions extraction needs cell location mapping'
            
        except Exception as e:
            results['error'] = str(e)
        
        return results
    
    def extract_input_parameters(self) -> Dict[str, Any]:
        """Extract input parameters from Expert Input or Main Input."""
        params = {}
        
        try:
            # Check Main Input & Output for configuration
            main_sheet = self.wb_values['Main Input & Output']
            
            # Configuration, capacity, etc. need to be mapped to specific cells
            params['note'] = 'Input parameter extraction needs cell location mapping'
            
        except Exception as e:
            params['error'] = str(e)
        
        return params
    
    def create_test_fixture(self, test_id: str, description: str = None) -> Dict[str, Any]:
        """
        Create a complete test fixture from current Excel state.
        
        Args:
            test_id: Test case ID (e.g., 'tc_001')
            description: Optional description
            
        Returns:
            Test fixture dictionary
        """
        fixture = {
            'test_id': test_id,
            'description': description or 'Extracted from Excel',
            'excel_file': os.path.basename(self.excel_path),
            'extraction_date': datetime.now().isoformat(),
            'metadata': {
                'extractor_version': '1.0',
                'excel_data_only': True,
                'note': 'Values extracted from last Excel calculation'
            },
            'inputs': {
                'assay': self.extract_current_assay(),
                'parameters': self.extract_input_parameters()
            },
            'expected_results': {
                'products': self.extract_product_yields(),
                'energy': self.extract_energy_data(),
                'emissions': self.extract_emissions_data()
            },
            'validation': {
                'relative_tolerance': 1e-5,
                'absolute_tolerance': 1e-8,
                'critical_fields': [
                    'products.gasoline_vol_pct',
                    'products.diesel_vol_pct',
                    'products.jet_vol_pct'
                ]
            }
        }
        
        return fixture
    
    def save_fixture(self, fixture: Dict[str, Any], output_path: str):
        """
        Save test fixture to JSON file.
        
        Args:
            fixture: Test fixture dictionary
            output_path: Path to output JSON file
        """
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(fixture, f, indent=2, default=str)
        
        print(f"✓ Test fixture saved to: {output_path}")
        print(f"  Test ID: {fixture['test_id']}")
        print(f"  Extracted: {fixture['extraction_date']}")


def extract_all_visible_data(excel_path: str) -> Dict[str, Any]:
    """
    Extract all visible numeric data from key sheets.
    
    Useful for initial exploration of what data is available.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    data = {}
    
    # Key sheets to extract from
    sheets_to_check = [
        'Results Single Assay',
        'Main Input & Output',
        'CokingRefinery Detailed Results'
    ]
    
    for sheet_name in sheets_to_check:
        if sheet_name not in wb.sheetnames:
            continue
        
        sheet = wb[sheet_name]
        sheet_data = {}
        
        # Extract first 100 rows, first 20 columns
        for row_idx in range(1, min(101, sheet.max_row + 1)):
            for col_idx in range(1, min(21, sheet.max_column + 1)):
                cell = sheet.cell(row_idx, col_idx)
                
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    coord = cell.coordinate
                    
                    # Try to get a label from nearby cells
                    label = None
                    for offset in [0, -1, -2]:
                        if col_idx + offset > 0:
                            label_cell = sheet.cell(row_idx, col_idx + offset)
                            if label_cell.value and isinstance(label_cell.value, str):
                                label = label_cell.value
                                break
                    
                    sheet_data[coord] = {
                        'value': cell.value,
                        'label': label
                    }
        
        data[sheet_name] = sheet_data
    
    return data


def main():
    """Main extraction function."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Extract reference data from PRELIM Excel file')
    parser.add_argument('--excel', default='PRELIM_v1.6.xlsm',
                       help='Path to PRELIM Excel file')
    parser.add_argument('--output', default='test_fixture.json',
                       help='Output JSON file path')
    parser.add_argument('--test-id', default='tc_001',
                       help='Test case ID')
    parser.add_argument('--description', default=None,
                       help='Test case description')
    parser.add_argument('--explore', action='store_true',
                       help='Explore mode: extract all visible data')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel):
        print(f"Error: Excel file not found: {args.excel}")
        return 1
    
    if args.explore:
        print("Exploring Excel file...")
        data = extract_all_visible_data(args.excel)
        explore_output = args.output.replace('.json', '_explore.json')
        with open(explore_output, 'w') as f:
            json.dump(data, f, indent=2, default=str)
        print(f"✓ Exploration data saved to: {explore_output}")
        return 0
    
    print(f"Extracting reference data from: {args.excel}")
    print(f"Test ID: {args.test_id}")
    
    extractor = ExcelReferenceExtractor(args.excel)
    fixture = extractor.create_test_fixture(args.test_id, args.description)
    extractor.save_fixture(fixture, args.output)
    
    # Print summary
    print("\nExtracted data summary:")
    print(f"  Products: {len([k for k in fixture['expected_results']['products'].keys() if 'error' not in k])}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())

