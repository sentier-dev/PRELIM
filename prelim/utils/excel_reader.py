"""
Utilities for reading and parsing Excel files.

Provides functions to extract data, formulas, and structure from PRELIM Excel files.
"""

import openpyxl
import re
from typing import Dict, Any, Optional, Tuple, List
from dataclasses import dataclass


@dataclass
class CellReference:
    """Represents a cell reference with sheet, row, and column."""
    sheet: Optional[str]
    row: int
    column: str
    
    def __str__(self):
        if self.sheet:
            return f"'{self.sheet}'!{self.column}{self.row}"
        return f"{self.column}{self.row}"
    
    @classmethod
    def parse(cls, ref: str) -> 'CellReference':
        """Parse a cell reference string like 'Sheet1'!A1 or B2."""
        if '!' in ref:
            parts = ref.split('!')
            sheet = parts[0].strip("'")
            cell_ref = parts[1]
        else:
            sheet = None
            cell_ref = ref
        
        # Extract column and row
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            col, row = match.groups()
            return cls(sheet=sheet, row=int(row), column=col)
        
        raise ValueError(f"Invalid cell reference: {ref}")


class ExcelFormulaExtractor:
    """Extracts formulas and data from Excel workbooks."""
    
    def __init__(self, filepath: str):
        """
        Initialize with path to Excel file.
        
        Args:
            filepath: Path to .xlsm or .xlsx file
        """
        self.filepath = filepath
        self._wb = None
        self._wb_values = None
    
    def load(self):
        """Load the workbook (with formulas and values)."""
        if self._wb is None:
            self._wb = openpyxl.load_workbook(self.filepath, data_only=False)
            self._wb_values = openpyxl.load_workbook(self.filepath, data_only=True)
    
    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names."""
        self.load()
        return self._wb.sheetnames
    
    def extract_cell_formula(self, sheet_name: str, cell_ref: str) -> Optional[str]:
        """
        Extract formula from a specific cell.
        
        Args:
            sheet_name: Name of the sheet
            cell_ref: Cell reference (e.g., 'A1', 'B10')
            
        Returns:
            Formula string or None if no formula
        """
        self.load()
        sheet = self._wb[sheet_name]
        cell = sheet[cell_ref]
        
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            return cell.value
        return None
    
    def extract_cell_value(self, sheet_name: str, cell_ref: str) -> Any:
        """
        Extract calculated value from a specific cell.
        
        Args:
            sheet_name: Name of the sheet
            cell_ref: Cell reference (e.g., 'A1', 'B10')
            
        Returns:
            Cell value
        """
        self.load()
        sheet = self._wb_values[sheet_name]
        return sheet[cell_ref].value
    
    def extract_sheet_formulas(self, sheet_name: str, 
                               max_rows: Optional[int] = None) -> Dict[str, str]:
        """
        Extract all formulas from a sheet.
        
        Args:
            sheet_name: Name of the sheet
            max_rows: Maximum number of rows to scan (None for all)
            
        Returns:
            Dictionary mapping cell coordinates to formula strings
        """
        self.load()
        sheet = self._wb[sheet_name]
        formulas = {}
        
        max_row = max_rows if max_rows else sheet.max_row
        
        for row in sheet.iter_rows(min_row=1, max_row=max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas[cell.coordinate] = cell.value
        
        return formulas
    
    def extract_dependencies(self, formula: str) -> List[CellReference]:
        """
        Extract cell references that a formula depends on.
        
        Args:
            formula: Formula string (starting with =)
            
        Returns:
            List of CellReference objects
        """
        dependencies = []
        
        # Pattern for sheet references like 'Sheet1'!A1 or Sheet1!A1
        pattern1 = r"'([^']+)'!([A-Z]+\d+)"
        matches1 = re.findall(pattern1, formula)
        for sheet, cell in matches1:
            try:
                ref = CellReference.parse(f"'{sheet}'!{cell}")
                dependencies.append(ref)
            except ValueError:
                pass
        
        # Pattern for same-sheet references like A1, B2
        # Exclude function names and ranges
        pattern2 = r'(?<![A-Z])([A-Z]+\d+)(?![A-Z0-9])'
        matches2 = re.findall(pattern2, formula)
        for cell in matches2:
            try:
                ref = CellReference.parse(cell)
                dependencies.append(ref)
            except ValueError:
                pass
        
        return dependencies
    
    def close(self):
        """Close the workbook."""
        if self._wb:
            self._wb.close()
            self._wb = None
        if self._wb_values:
            self._wb_values.close()
            self._wb_values = None

